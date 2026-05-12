from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.utils.dateparse import parse_date

from openpyxl import load_workbook

from myapp.models import (
    Budget, BudgetAllocation, EquipmentCategory, EquipmentItem,
    Project, ProjectChronology, ProjectEquipment, ProjectRequirement,
)
from myapp.risk import assess_project_delay_risk


def text(value):
    return str(value).strip() if value is not None else ''


def money(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    cleaned = text(value).replace('Php', '').replace('₱', '').replace(',', '').strip()
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def date_value(value):
    if not value:
        return None
    if hasattr(value, 'date'):
        return value.date()
    parsed = parse_date(text(value))
    return parsed


class Command(BaseCommand):
    help = 'Import the DOST Biliran GIA/CEST workbook into TaskPro project, budget, equipment, and checklist records.'

    def add_arguments(self, parser):
        parser.add_argument('workbook_path')

    def handle(self, *args, **options):
        workbook_path = options['workbook_path']
        try:
            wb = load_workbook(workbook_path, data_only=True)
        except Exception as exc:
            raise CommandError(f'Could not open workbook: {exc}')

        project_count = self._import_projects(wb)
        requirement_count = self._seed_requirements()
        equipment_count = self._import_equipment(wb)
        chronology_count = self._import_chronology(wb)

        for project in Project.objects.all():
            assess_project_delay_risk(project, save=True)

        self.stdout.write(self.style.SUCCESS(
            f'Import complete: {project_count} projects, {requirement_count} requirements, '
            f'{equipment_count} equipment records, {chronology_count} chronology entries.'
        ))

    def _import_projects(self, wb):
        if 'Projs 2022-2025' not in wb.sheetnames:
            return 0
        ws = wb['Projs 2022-2025']
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            code = text(row[1])
            title = text(row[4])
            if not code and not title:
                continue

            total_cost = money(row[13])
            released = money(row[11])
            budget = None
            fund_source = text(row[9])
            year = int(float(row[8])) if text(row[8]).replace('.', '', 1).isdigit() else None
            if fund_source and year:
                budget, _ = Budget.objects.get_or_create(
                    fiscal_year=year,
                    fund_source=fund_source,
                    defaults={'total_equipment_value': total_cost or Decimal('0.00')}
                )

            project, _ = Project.objects.update_or_create(
                project_code=code or None,
                project_title=title,
                defaults={
                    'no': int(float(row[0])) if text(row[0]).replace('.', '', 1).isdigit() else None,
                    'agency_grantee': text(row[2]),
                    'mun': text(row[3]),
                    'beneficiary': text(row[5]),
                    'beneficiary_address': text(row[6]),
                    'program': text(row[7]),
                    'year': year,
                    'fund_source': fund_source,
                    'funds': released or total_cost,
                    'total_project_cost': total_cost,
                    'total_funds_released': released,
                    'budget': budget,
                    'province': 'Biliran',
                    'district': 'Lone',
                }
            )
            count += 1

        if 'CEST Liquidation Status' in wb.sheetnames:
            ws = wb['CEST Liquidation Status']
            for row in ws.iter_rows(min_row=6, values_only=True):
                code = text(row[1])
                title = text(row[3])
                if not code and not title:
                    continue
                project = Project.objects.filter(project_code=code).first() or Project.objects.filter(project_title=title).first()
                if not project:
                    continue
                project.status_of_liquidation = text(row[7]) or project.status_of_liquidation
                project.remarks = text(row[8]) or project.remarks
                project.project_start = date_value(row[11]) or project.project_start
                project.project_end = date_value(row[12]) or project.project_end
                project.status = text(row[7]) or project.status
                project.save()
        return count

    def _seed_requirements(self):
        names = [
            'Letter of Intent to Avail the Assistance',
            'Endorsement of PSTDs/Project Leader',
            'TNA Report',
            'Project Proposal using GIA Proposal Format',
            'Project Line-Item-Budget',
            'SB Resolution/Board Resolution',
            'RTEC Report',
            'Compliance to iRTEC Comments',
            'iRTEC Endorsement with approval of RD',
            'Notarized MOA',
            'List of Beneficiaries',
            'Registration of Community Based Beneficiaries',
            'Terminal Report',
            'Liquidation Report',
        ]
        count = 0
        for project in Project.objects.all():
            for name in names:
                ProjectRequirement.objects.get_or_create(project=project, name=name)
                count += 1
        return count

    def _import_equipment(self, wb):
        if 'List of Equipment Purchased' not in wb.sheetnames:
            return 0
        ws = wb['List of Equipment Purchased']
        category, _ = EquipmentCategory.objects.get_or_create(name='GIA/CEST Equipment')
        count = 0
        current_project = None
        for row in ws.iter_rows(min_row=7, values_only=True):
            code = text(row[1])
            title = text(row[2])
            if code or title:
                current_project = Project.objects.filter(project_code=code).first() or Project.objects.filter(project_title=title).first()
                if not current_project and title:
                    current_project = Project.objects.create(
                        project_code=code or None,
                        project_title=title,
                        beneficiary=text(row[3]),
                        province='Biliran',
                        fund_source='GIA/CEST',
                        status='historical',
                    )
            if not current_project or not text(row[4]):
                continue

            unit_cost = money(row[9]) or Decimal('0.00')
            total_cost = money(row[10]) or unit_cost
            qty = int(float(row[7])) if text(row[7]).replace('.', '', 1).isdigit() else 1
            item, _ = EquipmentItem.objects.get_or_create(
                name=text(row[4]),
                category=category,
                defaults={
                    'unit': 'units',
                    'estimated_unit_cost': unit_cost,
                    'specifications': text(row[5]),
                }
            )
            budget = current_project.budget
            if not budget:
                budget, _ = Budget.objects.get_or_create(
                    fiscal_year=current_project.year or 2025,
                    fund_source=current_project.fund_source or 'GIA/CEST',
                )
                current_project.budget = budget
                current_project.save(update_fields=['budget', 'date_updated'])
            allocation, _ = BudgetAllocation.objects.get_or_create(
                budget=budget,
                equipment_item=item,
                defaults={'allocated_quantity': qty, 'delivered_quantity': qty, 'status': 'delivered'}
            )
            ProjectEquipment.objects.update_or_create(
                project=current_project,
                budget_allocation=allocation,
                serial_numbers=text(row[11]),
                defaults={
                    'delivered_quantity': qty,
                    'actual_specifications': text(row[6]),
                    'actual_unit_cost': unit_cost,
                    'actual_total_cost': total_cost,
                    'official_receipt': text(row[11]),
                    'supplier_name': text(row[12]),
                    'date_acquired': date_value(row[13]),
                }
            )
            count += 1
        return count

    def _import_chronology(self, wb):
        if 'Chronology of Events (2021 - on' not in wb.sheetnames:
            return 0
        ws = wb['Chronology of Events (2021 - on']
        count = 0
        project_blocks = []
        for col in range(1, ws.max_column + 1, 4):
            title = text(ws.cell(row=4, column=col).value)
            if title:
                project = Project.objects.filter(project_title__icontains=title[:80]).first()
                if project:
                    project_blocks.append((project, col))
        for project, col in project_blocks:
            for row in range(6, min(ws.max_row, 80) + 1):
                event = text(ws.cell(row=row, column=col + 1).value)
                if not event:
                    continue
                ProjectChronology.objects.get_or_create(
                    project=project,
                    event=event[:1000],
                    event_date=date_value(ws.cell(row=row, column=col).value),
                    defaults={'remarks': text(ws.cell(row=row, column=col + 2).value), 'source': 'excel'}
                )
                count += 1
        return count
